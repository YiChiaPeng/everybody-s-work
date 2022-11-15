# -*- coding:utf-8 -*-
import csv
import openpyxl
import time
import math
class Filter_args:
    def __init__(self):
        self.file_name="臺鐵每日各站分時OD資料(O).csv" #你要搜尋的檔案名稱
        #-----------------改這裡除了台北,板橋,一次放一到兩個就好--------------------
        self.originStationNames=["板橋"]#,"萬華","臺北","板橋","浮洲","樹林","南樹林","山佳","鶯歌","桃園","內壢","中壢"]
        self.needed_field=["TripDate","TripHour","OriginStationName","DestinationStationName","Volume"]
    
def filter(args):
    
    wbs=dict()
    counts=dict()
    for name in  args.originStationNames:
        wbs[name]=openpyxl.Workbook()
        counts[name]=0
        
    #create_sheet with station name as the id 
    
        
    
    # 開啟 CSV 檔案
    with open(args.file_name, newline='',encoding="utf-8") as csvfile:
        
        '''抓header'''
        list_rows=csv.reader(csvfile)
        i=0
        for list_row in list_rows:
            if i==0:
                chinese_header=list_row
                i+=1
            elif i==1:
                english_header=list_row
                i+=1
            else:
                break
        print(chinese_header)
        print(english_header)
        dic_english_to_chinese_header=dict(zip(english_header,chinese_header))
        dic_chinese_to_english_header=dict(zip(chinese_header,english_header))
        
        needed_header=[]
        for item in args.needed_field:
            needed_header.append(dic_english_to_chinese_header[item])
        
        
        
        # 讀取 CSV 檔案內容
        dict_rows = csv.DictReader(csvfile,fieldnames=english_header)
        for row in dict_rows:
            for name in args.originStationNames:
                if (row["OriginStationName"]==name):
                    
                    id =math.floor(counts[name]/1000000)
                    if counts[name]%1000000==0:
                        wbs[name].create_sheet(str(id))
                        wbs[name][str(id)].append(needed_header)
                    if(counts[name]==0):
                        wbs[name].remove(wbs[name]['Sheet'])
                    sheet=wbs[name].worksheets[id]
                    templist=[]
                    for item in args.needed_field:
                        templist.append(row[item])
                    sheet.append(templist)
                    counts[name]=counts[name]+1
                    
        for name in  args.originStationNames:
            wbs[name].save(str(name)+"臺鐵每日各站分時OD資料(O).xlsx")    
        

    
    print("excel files are generated")
    


    
    
if __name__=="__main__":
    start_time = time.time()
    args=Filter_args()
    filter(args)
    print("幹你伯公你跑了")
    print(str((time.time()-start_time)/60)+"分鐘")
    
            
